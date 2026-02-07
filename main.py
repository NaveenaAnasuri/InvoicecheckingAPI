from fastapi import FastAPI, File, UploadFile, HTTPException, BackgroundTasks
from fastapi.responses import FileResponse, JSONResponse
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, NamedStyle
import logging
import os
import signal
import time
from datetime import datetime

# -------------------------------------------------
# LOGGING SETUP
# -------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s"
)
logger = logging.getLogger(__name__)

app = FastAPI(title="Invoice Validation & Anomaly Detection API")

# -------------------------------------------------
# VALIDATION LOGIC
# -------------------------------------------------
def validate_invoice(row):
    issues = []

    try:
        if round(row["sell_price"] * row["qty"], 2) != round(row["total_sale_value"], 2):
            issues.append("Total sale mismatch")

        if round(row["cost_price"] * row["qty"], 2) != round(row["total_cost_value"], 2):
            issues.append("Total cost mismatch")

        if round(row["total_sale_value"] - row["total_cost_value"], 2) != round(row["profit"], 2):
            issues.append("Profit mismatch")

        if row["qty"] <= 0:
            issues.append("Invalid quantity")

    except Exception as e:
        issues.append(f"Validation error: {e}")

    return issues

# -------------------------------------------------
# GRACEFUL SHUTDOWN FUNCTION
# -------------------------------------------------
def shutdown_server():
    logger.info("Shutting down server automatically (CTRL+C equivalent)")
    time.sleep(2)  # allow response to complete
    os.kill(os.getpid(), signal.SIGINT)

# -------------------------------------------------
# API ENDPOINT
# -------------------------------------------------
@app.post("/upload_invoices")
async def upload_invoices(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    return_json: bool = False
):
    logger.info("Invoice upload started")

    if not file.filename.endswith((".json", ".csv")):
        raise HTTPException(status_code=400, detail="Only JSON or CSV files supported")

    content = await file.read()

    try:
        df = pd.read_json(BytesIO(content)) if file.filename.endswith(".json") else pd.read_csv(BytesIO(content))
        logger.info("File successfully read into DataFrame")
    except Exception as e:
        logger.error("File read failed")
        raise HTTPException(status_code=400, detail=f"File read error: {e}")

    # Duplicate detection
    df["is_duplicate"] = df.duplicated(subset=["invoice_number", "part_number"], keep=False)
    logger.info("Duplicate check completed")

    # Rule validation
    df["Issue Type"] = df.apply(validate_invoice, axis=1)
    logger.info("Rule-based validation completed")

    # Add duplicate flag
    df.loc[df["is_duplicate"], "Issue Type"] = df.loc[df["is_duplicate"], "Issue Type"].apply(
        lambda x: x + ["Duplicate invoice"]
    )

    # Final cleanup
    df["Issue Type"] = df["Issue Type"].apply(lambda x: ", ".join(x) if x else "")
    df["Status"] = df["Issue Type"].apply(lambda x: "Valid" if x == "" else "Invalid")
    df.drop(columns=["is_duplicate"], inplace=True)

    # Excel export with date in filename
    today_str = datetime.now().strftime("%m-%d-%Y")
    output_file = f"validated_{file.filename.split('.')[0]}_{today_str}.xlsx"
    df.to_excel(output_file, index=False)
    logger.info(f"Excel file generated: {output_file}")

    # Excel formatting
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    # Status color formatting
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    status_col = [cell.value for cell in ws[1]].index("Status") + 1

    # Date formatting (mm/dd/yyyy)
    if "date" in df.columns:
        date_style = NamedStyle(name="datetime", number_format='MM/DD/YYYY')
        date_col = [cell.value for cell in ws[1]].index("date") + 1
        for row in ws.iter_rows(min_row=2):
            if row[date_col - 1].value:
                row[date_col - 1].style = date_style

    # Apply status fill
    for row in ws.iter_rows(min_row=2):
        fill = red if row[status_col - 1].value == "Invalid" else green
        for cell in row:
            cell.fill = fill

    wb.save(output_file)

    logger.info("Validation completed successfully")

    # Auto shutdown after response
    background_tasks.add_task(shutdown_server)

    # Optional JSON return for ERP integration
    if return_json:
        return JSONResponse(content=df.to_dict(orient="records"))

    return FileResponse(
        output_file,
        filename=output_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.get("/")
def health_check():
    return {"status": "Invoice Validation API running"}
